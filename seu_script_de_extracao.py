import pdfplumber
import os
import re
import csv
import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# --- FUNÇÕES AUXILIARES (do seu script original) ---
# Todas as funções que seu amigo criou estão aqui, sem modificação.

def limpar_texto(texto):
    """Remove quebras de linha extras e espaços desnecessários."""
    if texto:
        return re.sub(r'\s+', ' ', str(texto)).strip()
    return ""

def extrair_dados_historico(caminho_pdf):
    """Extrai disciplinas pendentes e resumo de carga horária de um histórico PDF."""
    dados_pendentes = []  # lista de dicts: {codigo, nome, ch}
    resumo_horas = {"optativos": "0", "complementares": "0", "total": "0"}

    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                page_tem_pendentes = False  # Flag para saber se encontrou pendentes nesta página
                tables = page.extract_tables() or []
                
                for table in tables:
                    if not table or not table[0]:
                        continue

                    # --- Detecta tabela de carga horária integralizada/pendente ---
                    header_texto = " ".join(limpar_texto(c) for c in table[0] if c)
                    header_texto_up = header_texto.upper()
                    if ("CARGA" in header_texto_up and "HORÁRIA" in header_texto_up) or "OBRIGATÓRIAS" in header_texto_up:
                        for row in table:
                            if not row:
                                continue
                            primeira = limpar_texto(row[0]).upper() if row[0] else ""
                            if "PENDENTE" in primeira:
                                row_clean = [limpar_texto(c) for c in row if c is not None]
                                if len(row_clean) >= 4: # Ajustado para maior robustez
                                    resumo_horas["total"] = re.sub(r'[^0-9]', '', row_clean[-1]) or "0"
                                    resumo_horas["complementares"] = re.sub(r'[^0-9]', '', row_clean[-2]) or "0"
                                    resumo_horas["optativos"] = re.sub(r'[^0-9]', '', row_clean[-3]) or "0"
                                elif len(row_clean) == 3: # Caso não tenha a coluna "Obrigatórias"
                                    resumo_horas["total"] = re.sub(r'[^0-9]', '', row_clean[-1]) or "0"
                                    resumo_horas["complementares"] = re.sub(r'[^0-9]', '', row_clean[-2]) or "0"
                                    resumo_horas["optativos"] = re.sub(r'[^0-9]', '', row_clean[-3]) or "0"

                    # --- Detecta tabela de componentes curriculares obrigatórios pendentes ---
                    header = [limpar_texto(cell).upper() for cell in table[0] if cell]
                    header_found = ("CÓDIGO" in header and "COMPONENTE CURRICULAR" in header)
                    if not header_found:
                        continue

                    page_tem_pendentes = True
                    for row in table[1:]:
                        if not row or len(row) < 2:
                            continue
                        codigo = limpar_texto(row[0])
                        nome_disciplina = limpar_texto(row[1])
                        if not codigo or not nome_disciplina:
                            continue
                        if "ENADE" in codigo.upper() or "ENADE" in nome_disciplina.upper():
                            continue

                        ch = ""
                        for cell in row[2:]:
                            cell_text = limpar_texto(cell)
                            if re.match(r'^\d+[ ]*h?$', cell_text, flags=re.IGNORECASE):
                                numero = re.findall(r'\d+', cell_text)[0]
                                ch = f"{numero} h"
                                break

                        esta_matriculado = any("MATRICULADO" in limpar_texto(cell).upper() for cell in row if cell)
                        if esta_matriculado and "(Matriculado)" not in nome_disciplina:
                            nome_disciplina += " (Matriculado)"

                        dados_pendentes.append({"codigo": codigo, "nome": nome_disciplina, "ch": ch})
                
                # --- Fallback textual (Lógica original) ---
                if not page_tem_pendentes:
                    texto = page.extract_text() or ""
                    linhas = [limpar_texto(l) for l in texto.split('\n') if l.strip()]
                    capturando = False
                    for linha in linhas:
                        up = linha.upper()
                        if 'COMPONENTES CURRICULARES OBRIGATÓRIOS PENDENTES' in up:
                            capturando = True
                            continue
                        if capturando and any(p in up for p in ['INTEGRALIZADOS', 'SITUAÇÃO', 'CARGA HORÁRIA', 'TOTAL', 'OBSERVAÇÕES:', 'EQUIVALÊNCIAS:']):
                            capturando = False
                            break
                        if not capturando:
                            continue
                        
                        m = re.match(r'^(?P<codigo>[A-Z0-9]{6,})\s+(?P<resto>.+)$', linha)
                        if not m:
                            continue
                        codigo = m.group('codigo')
                        resto = m.group('resto')
                        m_ch = re.search(r'(\d+)\s*h$', resto)
                        ch = ''
                        if m_ch:
                            ch = f"{m_ch.group(1)} h"
                            nome = resto[:resto.rfind(m_ch.group(1))].strip()
                        else:
                            nome = resto
                        if 'ENADE' in nome.upper() or 'ENADE' in codigo.upper():
                            continue
                        
                        esta_matriculado = 'MATRICULADO' in nome.upper()
                        if esta_matriculado and '(MATRICULADO)' not in nome.upper():
                            nome += ' (Matriculado)'
                            
                        dados_pendentes.append({"codigo": codigo, "nome": nome, "ch": ch})
    except Exception as e:
        print(f"Erro ao ler o PDF {caminho_pdf}: {e}")
        return [], resumo_horas

    return dados_pendentes, resumo_horas

def gerar_resumo_string(dados_pendentes, resumo_horas):
    qtd = len(dados_pendentes)
    partes = []
    if qtd > 0:
        palavra_comp = "componente" if qtd == 1 else "componentes"
        partes.append(f"{qtd} {palavra_comp}")

    opt = resumo_horas.get('optativos', '0')
    compl = resumo_horas.get('complementares', '0')
    total = resumo_horas.get('total', '0')

    if opt and opt != '0':
        partes.append(f"{opt}h optativos")
    if compl and compl != '0':
        partes.append(f"complementares {compl}h")

    total_str = f"{total} h" if total else "0 h"
    prefixo = " + ".join(partes)
    if prefixo:
        return prefixo + f"; {total_str}"
    elif total and total != '0':
         return f"{total_str}"
    else:
        return "não contém" # Caso não tenha pendências

def extrair_matricula_do_nome_arquivo(nome_arquivo):
    match = re.search(r'historico[_-]?(\d+)', nome_arquivo, re.IGNORECASE)
    return match.group(1) if match else ""

def extrair_nome_aluno(caminho_pdf):
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            if not pdf.pages:
                return ""
            texto = pdf.pages[0].extract_text() or ""
            match = re.search(r'Nome:\s*([A-ZÀ-Ú\s]+?)(?:\s+Matrícula:|\s*$)', texto, re.MULTILINE)
            if match:
                return match.group(1).strip()
    except Exception as e:
        print(f"   Aviso: não foi possível extrair nome de {caminho_pdf}: {e}")
    return ""

def carregar_percentuais(arquivo_xls):
    percentuais = {}
    try:
        # Se nenhum arquivo fornecido, retorna dicionário vazio
        if not arquivo_xls:
            return percentuais

        # Suporta .xls (xlrd) e .xlsx (openpyxl). Detecta pela extensão do arquivo.
        if not os.path.exists(arquivo_xls):
            print(f"   Aviso: arquivo de percentuais não encontrado: {arquivo_xls}")
            return percentuais

        _, ext = os.path.splitext(arquivo_xls)
        ext = ext.lower()

        if ext == '.xls':
            wb = xlrd.open_workbook(arquivo_xls)
            ws = wb.sheet_by_index(0)
            for row_idx in range(9, ws.nrows):
                row = ws.row_values(row_idx)
                if len(row) > 6:
                    matricula_val = row[1]  # Coluna B (MATRICULA)
                    percentual_val = row[6]  # Coluna G (PERCENTUAL CUMPRIDO)
                    if isinstance(matricula_val, float):
                        matricula = str(int(matricula_val))
                    else:
                        matricula = str(matricula_val).strip()
                    percentual = str(percentual_val).strip() if percentual_val else ""
                    if matricula and percentual:
                        percentuais[matricula] = percentual

        elif ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
            # Usa openpyxl para arquivos xlsx
            wb = load_workbook(arquivo_xls, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            # openpyxl rows are 1-indexed; dados começam na linha 10 (índice humano)
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx < 10:
                    continue
                # row é uma tupla de valores
                row_vals = list(row)
                if len(row_vals) > 6:
                    matricula_val = row_vals[1]
                    percentual_val = row_vals[6]
                    if isinstance(matricula_val, float):
                        matricula = str(int(matricula_val))
                    else:
                        matricula = str(matricula_val).strip()
                    percentual = str(percentual_val).strip() if percentual_val else ""
                    if matricula and percentual:
                        percentuais[matricula] = percentual

        else:
            print(f"   Aviso: formato de arquivo não suportado para percentuais: {arquivo_xls}")

    except Exception as e:
        print(f"   Aviso: não foi possível carregar percentuais de {arquivo_xls}: {e}")
    return percentuais


# --- FUNÇÃO PRINCIPAL ADAPTADA ---
# Esta é a função que o app.py irá chamar.

def run_extraction_process_web_mode(pdf_upload_folder, excel_percentual_path, output_report_folder, progress_callback=None):
    """
    Executa o processo de extração principal.
    Recebe os caminhos das pastas (do servidor) e gera os relatórios.
    Retorna um dicionário com os nomes dos arquivos gerados.
    progress_callback: função opcional que recebe (current, total) para reportar progresso
    """
    
    # 1. Carrega percentuais (se informado)
    if excel_percentual_path:
        print(f"Carregando percentuais de '{excel_percentual_path}'...")
        percentuais_dict = carregar_percentuais(excel_percentual_path)
        print(f"   → {len(percentuais_dict)} percentuais carregados.\n")
    else:
        print("Nenhum arquivo de percentuais fornecido — extração seguirá sem percentuais.")
        percentuais_dict = {}
    
    # 2. Lista os PDFs (da pasta de upload)
    pdfs_encontrados = sorted([f for f in os.listdir(pdf_upload_folder) if f.lower().endswith(".pdf")])
    print(f"Encontrados {len(pdfs_encontrados)} arquivos PDF na pasta de upload. Iniciando extração...")

    # 3. Define os nomes dos arquivos de saída
    excel_output_name = "relatorio_componentes.xlsx"
    csv_compact_name = "relatorio_final.csv"
    txt_output_name = "relatorio_historicos.txt"
    
    # Define os caminhos completos de saída
    excel_output_path = os.path.join(output_report_folder, excel_output_name)
    csv_compact_path = os.path.join(output_report_folder, csv_compact_name)
    txt_output_path = os.path.join(output_report_folder, txt_output_name)

    # 4. Cria o Workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Componentes Pendentes"

    headers = [None, 'Matrícula', 'Nome', 'E-mail', 'Componentes Pendentes', 
               'Quantidade de \n Componentes', 'CH Pendente', 'Percentual\nCumprido']
    ws.append(headers)
    
    for cell in ws[1]:
        if cell.value:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    seq = 1  # Contador sequencial
    
    # 5. Abre os arquivos de saída (CSV e TXT) e processa os PDFs
    with open(csv_compact_path, "w", newline='', encoding="utf-8-sig") as csv_compact, \
         open(txt_output_path, "w", encoding="utf-8-sig") as arquivo_txt:

        writer_compacto = csv.writer(csv_compact, delimiter=';')
        writer_compacto.writerow(['Linha Consolidada','Arquivo'])

        for i, arquivo in enumerate(pdfs_encontrados):
            print(f"Processando [{i+1}/{len(pdfs_encontrados)}]: {arquivo}")
            
            # Reporta progresso se callback fornecido
            if progress_callback:
                progress_callback(i + 1, len(pdfs_encontrados))
            
            caminho_completo = os.path.join(pdf_upload_folder, arquivo)
            
            # --- Extrai os dados do PDF ---
            pendentes, resumo = extrair_dados_historico(caminho_completo)
            
            matricula = extrair_matricula_do_nome_arquivo(arquivo)
            nome_aluno = extrair_nome_aluno(caminho_completo)
            percentual = percentuais_dict.get(matricula, "")
            resumo_qtd = gerar_resumo_string(pendentes, resumo)
            ch_total = f"{resumo.get('total','0')} h"

            if not pendentes:
                linha_consolidada = f"não contém ; {resumo_qtd}"
                writer_compacto.writerow([linha_consolidada, arquivo])
                arquivo_txt.write(linha_consolidada + "\n")
                
                ws.append([seq, matricula, nome_aluno, None, 'não contém', resumo_qtd, ch_total, percentual])
                seq += 1
                continue

            # Com disciplinas pendentes
            for idx, d in enumerate(pendentes):
                codigo = d.get('codigo','')
                nome = d.get('nome','')
                ch = d.get('ch','')
                componente_texto = f"{codigo} {nome} {ch}".strip()
                
                if idx == 0:
                    linha_consolidada = f"{componente_texto} ; {resumo_qtd}".strip()
                    writer_compacto.writerow([linha_consolidada, arquivo])
                    arquivo_txt.write(linha_consolidada + "\n")
                    
                    ws.append([seq, matricula, nome_aluno, None, componente_texto, resumo_qtd, ch_total, percentual])
                    seq += 1
                else:
                    writer_compacto.writerow([componente_texto, arquivo])
                    arquivo_txt.write(componente_texto + "\n")
                    # Repetir matrícula, nome (e percentual quando disponível) em todas as linhas
                    ws.append([None, matricula, nome_aluno, None, componente_texto, None, None, None])

    # 6. Salva o Excel
    wb.save(excel_output_path)
    
    print("\nProcessamento concluído! Arquivos gerados em 'generated_reports'.")
    
    # 7. Retorna os nomes dos arquivos para o Flask
    return {
        'excel_report': excel_output_name,
        'csv_report': csv_compact_name,
        'txt_report': txt_output_name
    }